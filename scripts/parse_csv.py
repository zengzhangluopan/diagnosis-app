#!/usr/bin/env python3
"""
生意参谋文件解析 - 入口脚本 V2.0
更新: 2026-06-11

支持解析用户上传的生意参谋导出文件：
1. 商品概况 (.csv/.xls) — 按天的商品核心指标
2. 流量二级来源 (.csv/.xls) — 分渠道访客/转化数据
3. 推广报表 (.xlsx) — 阿里妈妈推广计划数据（由ad_parser.py处理）

输出: JSON格式，可直接传入 run_diagnosis.py 使用

用法:
  python parse_csv.py '<文件路径1>' '<文件路径2>' ...
  支持传入多个文件路径，自动识别文件类型并合并解析
"""

import sys
import json
import os
import re

# 添加父目录到sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def safe_float(value, default=0.0):
    """安全转浮点数，处理逗号、百分号、横杠等格式"""
    if value is None or value == '' or value == '-':
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(',', '').replace('，', '').replace('%', '').replace('％', '').strip()
    if not s or s == '-':
        return default
    try:
        return float(s)
    except ValueError:
        return default


def safe_int(value, default=0):
    """安全转整数"""
    f = safe_float(value, default)
    return int(f) if f != default else default


def detect_file_type(filepath: str) -> str:
    """根据文件名关键词判断文件类型"""
    name = os.path.basename(filepath).lower()

    # 推广报表：阿里妈妈导出，包含"计划报表"
    if '计划报表' in name or '推广' in name:
        return 'ad_report'

    # 流量来源：包含"流量"或"来源"
    if '流量' in name or '来源' in name:
        return 'traffic_source'

    # 商品概况：包含"商品"或默认
    if '商品' in name:
        return 'product_overview'

    # 无法判断时，尝试按内容判断
    return 'unknown'


def read_xls(filepath: str) -> tuple:
    """
    读取.xls格式文件（生意参谋导出的.csv实际是.xls）
    返回 (header_row, data_rows, all_rows)
    """
    try:
        import xlrd
    except ImportError:
        return None, [], []

    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
    except Exception:
        return None, [], []

    all_rows = []
    for r in range(ws.nrows):
        row = [ws.cell_value(r, c) for c in range(ws.ncols)]
        all_rows.append(row)

    # 找到表头行：第一个非空行，且包含中文字段名
    header_idx = None
    for i, row in enumerate(all_rows):
        # 表头行特征：至少5个非空列，且包含中文
        non_empty = [v for v in row if str(v).strip()]
        if len(non_empty) >= 5:
            has_chinese = any(re.search(r'[\u4e00-\u9fff]', str(v)) for v in non_empty)
            if has_chinese:
                header_idx = i
                break

    if header_idx is None:
        return None, [], all_rows

    header = [str(v).strip() for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    return header, data_rows, all_rows


def read_xlsx(filepath: str) -> tuple:
    """读取.xlsx格式文件"""
    try:
        import openpyxl
    except ImportError:
        return None, [], []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
    except Exception:
        return None, [], []

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    # 找表头行
    header_idx = None
    for i, row in enumerate(all_rows):
        non_empty = [v for v in row if v and str(v).strip()]
        if len(non_empty) >= 5:
            has_chinese = any(re.search(r'[\u4e00-\u9fff]', str(v)) for v in non_empty)
            if has_chinese:
                header_idx = i
                break

    if header_idx is None:
        return None, [], all_rows

    header = [str(v).strip() if v else '' for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    return header, data_rows, all_rows


def read_file(filepath: str) -> tuple:
    """根据扩展名选择读取方式，兼容.csv实际是.xls的情况"""
    ext = os.path.splitext(filepath)[1].lower()

    # 优先尝试xlrd（兼容.csv实际为.xls的情况）
    if ext in ('.csv', '.xls'):
        header, data, all_rows = read_xls(filepath)
        if header is not None:
            return header, data, all_rows
        # xlrd失败，尝试openpyxl
        return read_xlsx(filepath)

    if ext == '.xlsx':
        return read_xlsx(filepath)

    # 未知扩展名，两种都试
    header, data, all_rows = read_xls(filepath)
    if header is not None:
        return header, data, all_rows
    return read_xlsx(filepath)


def find_col(header: list, keywords: list) -> int:
    """在表头中查找包含关键词的列，返回列索引，找不到返回-1"""
    for kw in keywords:
        for i, h in enumerate(header):
            if kw in str(h):
                return i
    return -1


def parse_product_overview(filepath: str) -> dict:
    """
    解析商品概况文件
    提取：日均访客、整体转化率、加购率、跳失率、退款金额、支付件数等
    """
    header, data_rows, all_rows = read_file(filepath)
    if header is None:
        return {'error': f'无法解析商品概况文件: {filepath}'}

    # 找关键列
    col_visitors = find_col(header, ['商品访客数', '访客数'])
    col_conv_rate = find_col(header, ['支付转换率', '支付转化率'])
    col_bounce_rate = find_col(header, ['跳出率', '跳失率'])
    col_cart_persons = find_col(header, ['商品加购人数', '加购人数'])
    col_pay_amount = find_col(header, ['支付金额'])
    col_pay_pieces = find_col(header, ['支付件数'])
    col_pay_buyers = find_col(header, ['支付买家数'])
    col_refund = find_col(header, ['成功退款金额', '退款金额'])
    col_views = find_col(header, ['商品浏览量', '浏览量'])
    col_stay = find_col(header, ['平均停留时长', '停留时长'])
    col_collect = find_col(header, ['商品收藏人数', '收藏人数'])

    # 汇总每日数据
    total_visitors = 0
    total_pay_amount = 0
    total_pay_pieces = 0
    total_pay_buyers = 0
    total_refund = 0
    total_cart = 0
    total_collect = 0
    total_views = 0
    bounce_rates = []
    conv_rates = []
    stay_durations = []
    days = 0

    for row in data_rows:
        visitors = safe_float(row[col_visitors] if col_visitors >= 0 else 0)
        if visitors <= 0:
            continue  # 跳过空行

        days += 1
        total_visitors += visitors
        total_pay_amount += safe_float(row[col_pay_amount] if col_pay_amount >= 0 else 0)
        total_pay_pieces += safe_float(row[col_pay_pieces] if col_pay_pieces >= 0 else 0)
        total_pay_buyers += safe_float(row[col_pay_buyers] if col_pay_buyers >= 0 else 0)
        total_refund += safe_float(row[col_refund] if col_refund >= 0 else 0)
        total_cart += safe_float(row[col_cart_persons] if col_cart_persons >= 0 else 0)
        total_collect += safe_float(row[col_collect] if col_collect >= 0 else 0)
        total_views += safe_float(row[col_views] if col_views >= 0 else 0)

        if col_bounce_rate >= 0:
            br = safe_float(row[col_bounce_rate])
            if br > 0:
                bounce_rates.append(br)
        if col_conv_rate >= 0:
            cr = safe_float(row[col_conv_rate])
            if cr > 0:
                conv_rates.append(cr)
        if col_stay >= 0:
            sd = safe_float(row[col_stay])
            if sd > 0:
                stay_durations.append(sd)

    if days == 0:
        return {'error': '商品概况文件中未找到有效数据行'}

    # 计算汇总指标
    result = {
        'daily_visitors': round(total_visitors / days, 0),
        'conv_rate': round(total_pay_buyers / total_visitors * 100, 2) if total_visitors > 0 else 0,
        'bounce_rate': round(sum(bounce_rates) / len(bounce_rates), 2) if bounce_rates else None,
        'cart_rate': round(total_cart / total_visitors * 100, 2) if total_visitors > 0 else 0,
        'total_cart': total_cart,
        'collect_rate': round(total_collect / total_visitors * 100, 2) if total_visitors > 0 else 0,
        'monthly_sales': safe_int(total_pay_pieces),
        'total_pay_amount': round(total_pay_amount, 2),
        'total_refund_amount': round(total_refund, 2),
        'refund_rate': round(total_refund / total_pay_amount * 100, 2) if total_pay_amount > 0 else 0,
        'avg_stay_duration': round(sum(stay_durations) / len(stay_durations), 0) if stay_durations else None,
        'date_range_days': days,
    }

    # 移除None值
    result = {k: v for k, v in result.items() if v is not None}

    return result


def parse_traffic_source(filepath: str) -> dict:
    """
    解析流量二级来源文件
    提取：广告流量占比、搜索转化率、广告转化率、各渠道访客数等
    """
    header, data_rows, all_rows = read_file(filepath)
    if header is None:
        return {'error': f'无法解析流量来源文件: {filepath}'}

    # 找关键列
    col_l1 = find_col(header, ['一级来源'])
    col_l2 = find_col(header, ['二级来源'])
    col_l3 = find_col(header, ['三级来源'])
    col_l4 = find_col(header, ['四级来源'])
    col_visitors = find_col(header, ['访客数'])
    col_conv_rate = find_col(header, ['支付转化率', '支付转换率'])
    col_pay_buyers = find_col(header, ['支付买家数'])
    col_pay_amount = find_col(header, ['支付金额'])
    col_cart = find_col(header, ['加购人数'])
    col_collect = find_col(header, ['商品收藏人数', '收藏人数'])

    if col_visitors < 0:
        return {'error': '流量来源文件中未找到"访客数"列'}

    # 按二级来源汇总（生意参谋结构：一级来源 > 二级来源 > 三级/四级）
    # 关键：搜索/推荐/直播等是二级来源，挂在"经营优势"下面
    l2_data = {}  # {一级来源|二级来源: {visitors, pay_buyers, pay_amount, cart, collect}}
    l1_summary = {}  # 一级来源汇总

    cart_fav_visitors_direct = 0  # L3/L4购物车直接累加
    for row in data_rows:
        l1 = str(row[col_l1]).strip() if col_l1 >= 0 and col_l1 < len(row) else ''
        l2 = str(row[col_l2]).strip() if col_l2 >= 0 and col_l2 < len(row) else ''
        l3 = str(row[col_l3]).strip() if col_l3 >= 0 and col_l3 < len(row) else ''
        l4 = str(row[col_l4]).strip() if col_l4 >= 0 and col_l4 < len(row) else ''
        if l1:
            pass  # l1 used below

        visitors = safe_float(row[col_visitors] if col_visitors >= 0 and col_visitors < len(row) else 0)
        if visitors <= 0:
            continue

        buyers = safe_float(row[col_pay_buyers] if col_pay_buyers >= 0 else 0)
        amount = safe_float(row[col_pay_amount] if col_pay_amount >= 0 else 0)
        carts = safe_float(row[col_cart] if col_cart >= 0 else 0)
        collects = safe_float(row[col_collect] if col_collect >= 0 else 0)

        # 一级来源汇总行（二级=汇总 或 一级独立渠道）
        if l2 == '汇总' or (not l2 and l1):
            if l1 not in l1_summary:
                l1_summary[l1] = {'visitors': 0, 'pay_buyers': 0, 'pay_amount': 0, 'cart': 0, 'collect': 0}
            l1_summary[l1]['visitors'] += visitors
            l1_summary[l1]['pay_buyers'] += buyers
            l1_summary[l1]['pay_amount'] += amount
            l1_summary[l1]['cart'] += carts
            l1_summary[l1]['collect'] += collects

        # 二级来源非汇总行（这是真正的渠道细分）
        if l2 and l2 != '汇总':
            # 只取二级来源自身汇总（三级=同名 或 三级=汇总）
            if l3 == l2 or l3 == '汇总' or (l3 == '' and l4 == ''):
                key = f'{l1}|{l2}'
                if key not in l2_data:
                    l2_data[key] = {'l1': l1, 'l2': l2, 'visitors': 0, 'pay_buyers': 0, 'pay_amount': 0, 'cart': 0, 'collect': 0}
                l2_data[key]['visitors'] += visitors
                l2_data[key]['pay_buyers'] += buyers
                l2_data[key]['pay_amount'] += amount
                l2_data[key]['cart'] += carts
                l2_data[key]['pay_amount'] += amount
                l2_data[key]['collect'] += collects

        # 购物车渠道：L3或L4含"购物车"时直接累加（不受l2_data入口限制）
        if '购物车' in l3 or '购物车' in l4:
            cart_fav_visitors_direct += visitors

    if not l1_summary and not l2_data:
        return {'error': '流量来源文件中未找到有效数据'}

    # 计算总访客数（从一级来源汇总算）
    total_visitors = sum(d['visitors'] for d in l1_summary.values())

    # 按渠道分类：优先用二级来源细分，L1汇总兜底
    paid_visitors = 0
    search_visitors = 0
    recommend_visitors = 0
    content_visitors = 0
    cart_fav_visitors = 0
    other_visitors = 0

    search_data = {'visitors': 0, 'pay_buyers': 0, 'cart': 0, 'pay_amount': 0}
    paid_data = {'visitors': 0, 'pay_buyers': 0, 'cart': 0, 'pay_amount': 0}

    # 把L3/L4购物车直接累加结果合并
    cart_fav_visitors += cart_fav_visitors_direct

    # 先从二级来源细分统计
    for key, data in l2_data.items():
        l1, l2 = data['l1'], data['l2']
        v = data['visitors']

        if '付费' in l1 or '推广' in l1:
            paid_visitors += v
            paid_data['visitors'] += v
            paid_data['pay_buyers'] += data['pay_buyers']
            paid_data['cart'] += data['cart']
            paid_data['pay_amount'] += data.get('pay_amount', 0)
        elif '搜索' in l2:
            search_visitors += v
            search_data['visitors'] += v
            search_data['pay_buyers'] += data['pay_buyers']
            search_data['cart'] += data['cart']
            search_data['pay_amount'] += data.get('pay_amount', 0)
        elif '推荐' in l2 or '猜你' in l2:
            recommend_visitors += v
        elif '直播' in l2 or '短视频' in l2 or '内容' in l2 or '逛逛' in l2:
            content_visitors += v
        # 购物车已由L3/L4直接累加处理，此处不再重复
        elif '店内' in l1 or '导流' in l2:
            other_visitors += v  # 店内流转算其他
        elif '回访' in l1 or '主动' in l1:
            other_visitors += v  # 主动回访不计入购物车渠道
        else:
            other_visitors += v

    # 如果二级来源没覆盖到的L1，用L1汇总兜底
    l2_covered_l1 = set(data['l1'] for data in l2_data.values())
    for l1, data in l1_summary.items():
        if l1 not in l2_covered_l1:
            v = data['visitors']
            if '付费' in l1 or '推广' in l1:
                paid_visitors += v
                paid_data['visitors'] += v
                paid_data['pay_buyers'] += data['pay_buyers']
                paid_data['cart'] += data['cart']
            elif '搜索' in l1:
                search_visitors += v
                search_data['visitors'] += v
                search_data['pay_buyers'] += data['pay_buyers']
                search_data['cart'] += data['cart']
            elif '推荐' in l1:
                recommend_visitors += v
            elif '店内' in l1 or '流转' in l1:
                other_visitors += v
            # 购物车已由L3/L4直接累加处理，此处不再重复
            elif '回访' in l1 or '主动' in l1:
                other_visitors += v
            else:
                other_visitors += v

    search_conv_rate = round(search_data['pay_buyers'] / search_data['visitors'] * 100, 2) if search_data['visitors'] > 0 else None
    paid_conv_rate = round(paid_data['pay_buyers'] / paid_data['visitors'] * 100, 2) if paid_data['visitors'] > 0 else None

    # 分渠道加购率（人货匹配核心数据）
    search_cart_rate = round(search_data['cart'] / search_data['visitors'] * 100, 2) if search_data['visitors'] > 0 else None
    paid_cart_rate = round(paid_data['cart'] / paid_data['visitors'] * 100, 2) if paid_data['visitors'] > 0 else None

    result = {
        'ad_traffic_ratio': round(paid_visitors / total_visitors, 4) if total_visitors > 0 else 0,
        'natural_conv_rate': search_conv_rate,
        'ad_conv_rate': paid_conv_rate,
        'search_cart_rate': search_cart_rate,
        'paid_cart_rate': paid_cart_rate,
        'search_traffic_ratio': round(search_visitors / total_visitors, 4) if total_visitors > 0 else 0,
        'recommend_traffic_ratio': round(recommend_visitors / total_visitors, 4) if total_visitors > 0 else 0,
        'content_traffic_ratio': round(content_visitors / total_visitors, 4) if total_visitors > 0 else 0,
        'cart_fav_traffic_ratio': round(cart_fav_visitors / total_visitors, 4) if total_visitors > 0 else 0,
        'paid_visitors': paid_visitors,
        'search_visitors': search_visitors,
        'recommend_visitors': recommend_visitors,
        'content_visitors': content_visitors,
        'cart_fav_visitors': cart_fav_visitors,
        'search_pay_amount': round(search_data.get('pay_amount', 0), 2) if search_data.get('pay_amount') else None,
        'search_pay_buyers': safe_int(search_data.get('pay_buyers', 0)) if search_data.get('pay_buyers') else None,
        'paid_pay_amount': round(paid_data.get('pay_amount', 0), 2) if paid_data.get('pay_amount') else None,
        'paid_pay_buyers': safe_int(paid_data.get('pay_buyers', 0)) if paid_data.get('pay_buyers') else None,
    }

    # 移除None值
    result = {k: v for k, v in result.items() if v is not None}

    return result


def parse_files(file_paths: list) -> dict:
    """
    解析多个文件，自动识别类型，合并输出
    返回可直接传入 run_diagnosis.py 的 JSON 数据
    """
    result = {}

    for filepath in file_paths:
        if not os.path.exists(filepath):
            result['_errors'] = result.get('_errors', [])
            result['_errors'].append(f'文件不存在: {filepath}')
            continue

        file_type = detect_file_type(filepath)

        if file_type == 'product_overview':
            parsed = parse_product_overview(filepath)
            if 'error' in parsed:
                result['_errors'] = result.get('_errors', [])
                result['_errors'].append(parsed['error'])
            else:
                result.update(parsed)

        elif file_type == 'traffic_source':
            parsed = parse_traffic_source(filepath)
            if 'error' in parsed:
                result['_errors'] = result.get('_errors', [])
                result['_errors'].append(parsed['error'])
            else:
                result.update(parsed)

        elif file_type == 'ad_report':
            # 推广报表路径传递给 run_diagnosis.py，由 ad_parser.py 处理
            result['ad_report_path'] = filepath

        else:
            # 未知类型，尝试按内容判断
            header, data_rows, _ = read_file(filepath)
            if header:
                header_str = ' '.join(str(h) for h in header if h)
                if '访客数' in header_str and '支付转换率' in header_str and '跳出率' in header_str:
                    parsed = parse_product_overview(filepath)
                    if 'error' not in parsed:
                        result.update(parsed)
                elif '一级来源' in header_str and '访客数' in header_str:
                    parsed = parse_traffic_source(filepath)
                    if 'error' not in parsed:
                        result.update(parsed)
                else:
                    result['_errors'] = result.get('_errors', [])
                    result['_errors'].append(f'无法识别文件类型: {filepath}')

    # 自动识别类目（从推广报表计划名或文件名中提取关键词）
    category = _detect_category(result, file_paths)
    if category:
        result['category'] = category

    # 自动提取商品前缀（从推广报表计划名中提取）
    prefixes = _detect_product_prefix(file_paths)
    if prefixes:
        if len(prefixes) == 1:
            # 只有一个商品，自动选用
            result['product_prefix'] = list(prefixes.keys())[0]
        else:
            # 多个商品，列出所有前缀供选择
            result['_available_prefixes'] = prefixes
            # 暂不设product_prefix，等用户确认后补充

    # 自动提取商品名（从文件名中提取商品ID）
    product_name = _detect_product_name(file_paths)
    if product_name and 'product_name' not in result:
        result['product_name'] = product_name

    # 清理内部标记
    if '_l1_detail' in result:
        del result['_l1_detail']

    return result


# ============================================================
# 类目/商品自动识别
# ============================================================

# 品类关键词 → 类目映射
CATEGORY_KEYWORDS = {
    '净水器': ['净水器', '净水机', '净水壶', '滤芯', '直饮', '反渗透', '超滤'],
    '女装': ['女装', '连衣裙', '半身裙', '衬衫', '外套', '卫衣', '针织'],
    '食品': ['食品', '零食', '坚果', '茶叶', '干货', '糕点', '特产'],
    '3C数码': ['手机', '电脑', '平板', '耳机', '充电', '数码', '智能手表', '键盘'],
    '美妆': ['美妆', '护肤', '面膜', '精华', '口红', '粉底', '防晒', '彩妆'],
    '家居日用': ['家居', '日用', '收纳', '清洁', '毛巾', '枕头', '抱枕', '锅具'],
    '母婴': ['母婴', '奶粉', '纸尿裤', '婴儿', '童装', '孕妈', '辅食'],
}


def _detect_category(result: dict, file_paths: list) -> str:
    """
    自动识别类目，优先级：
    1. 推广报表计划名中的品类关键词
    2. 文件名中的品类关键词
    3. 基于客单价推断
    4. 返回空（由Bot或默认值处理）
    """
    all_text = ''

    # 从推广报表中读取计划名
    for filepath in file_paths:
        if detect_file_type(filepath) == 'ad_report':
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
                    if row and len(row) > 6 and row[6]:
                        all_text += str(row[6]) + ' '
                wb.close()
            except Exception:
                pass

    # 从文件名收集
    for fp in file_paths:
        all_text += os.path.basename(fp) + ' '

    # 匹配品类关键词
    best_category = ''
    best_count = 0
    for category, keywords in CATEGORY_KEYWORDS.items():
        count = sum(all_text.count(kw) for kw in keywords)
        if count > best_count:
            best_count = count
            best_category = category

    if best_count >= 2:  # 至少匹配2次才采纳
        return best_category

    # 基于客单价推断
    total_pay = result.get('total_pay_amount', 0)
    total_pieces = result.get('monthly_sales', 1)
    if total_pay > 0 and total_pieces > 0:
        avg_price = total_pay / total_pieces
        if avg_price >= 200:
            return '净水器'  # 高客单价默认净水器（用户主品类）
        elif avg_price >= 100:
            return '3C数码'
        elif avg_price >= 50:
            return '美妆'

    return ''


def _detect_product_prefix(file_paths: list) -> dict:
    """
    从推广报表计划名中提取所有商品前缀（如 CB073、CG104、A501）
    返回 {prefix: plan_count}，由调用方决定选哪个
    如果只有一个前缀，自动选用；多个前缀时需用户确认
    """
    prefixes = {}

    for filepath in file_paths:
        if detect_file_type(filepath) != 'ad_report':
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
                if row and len(row) > 6 and row[6]:
                    plan_name = str(row[6])
                    # 提取计划名开头的字母+数字前缀
                    match = re.match(r'^([A-Z]+\d+)', plan_name)
                    if match:
                        prefix = match.group(1)
                        prefixes[prefix] = prefixes.get(prefix, 0) + 1
            wb.close()
        except Exception:
            pass

    return prefixes


def _detect_product_name(file_paths: list) -> str:
    """从商品概况文件名中提取商品ID"""
    for fp in file_paths:
        name = os.path.basename(fp)
        # 匹配 "商品-XXXXXXXX" 格式
        match = re.search(r'商品[-_](\d+)', name)
        if match:
            return f'商品{match.group(1)}'
    return ''


def main():
    if len(sys.argv) < 2:
        print(json.dumps({
            'error': '请提供文件路径',
            'usage': 'python parse_csv.py <文件路径1> [文件路径2] ...'
        }, ensure_ascii=False))
        sys.exit(1)

    file_paths = sys.argv[1:]

    try:
        result = parse_files(file_paths)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        print(json.dumps({
            'error': f'文件解析失败: {str(e)}'
        }, ensure_ascii=False))
        sys.exit(1)


if __name__ == '__main__':
    main()
