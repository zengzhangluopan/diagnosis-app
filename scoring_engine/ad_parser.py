"""
推广报表解析器 - V6新增核心模块
版本: V1.0
更新: 2026-06-10

V6核心升级：
  1. 解析阿里妈妈推广报表（xlsx），按商品前缀筛选计划
  2. 拉新/收割自动分层（基于计划名关键词）
  3. 归因修正：剥离"自然流量转化金额"，计算ROI
  4. 成交新客占比正确解读（不用于判断是否在拉新）
  5. 退款/秒退分析（区分产品问题vs流量不精准）

输入: 阿里妈妈推广报表xlsx文件路径 + 商品名前缀
输出: 结构化推广诊断数据（供结论生成器使用）
"""

import os
import json
from typing import Optional
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# ============================================================
# 一、拉新/收割分类关键词
# ============================================================

LAUNCH_KEYWORDS = ['拉新', '精准', '关键词', '自定义', '短视频', '内容', '测款', '测图', '广泛']
HARVEST_KEYWORDS = ['收割', '重定向', '好货快投', '回访', '老客', '忠诚', '金卡', '高购买', '人群资产', '获取成交量']

# 全站推广特殊处理——拉新收割混合型，AI自动分配比例
# 默认拉新:收割 = 1:1，可在product_context中通过hybrid_ratio覆盖
HYBRID_PLANS = ['全站']
# 混合型计划的默认拉新/收割花费拆分比例
DEFAULT_HYBRID_LAUNCH_RATIO = 0.5  # 50%算拉新，50%算收割


def classify_plan(plan_name: str) -> str:
    """根据计划名判断拉新/收割类型
    
    规则：
    - 包含收割关键词 → harvest
    - 包含拉新关键词 → launch
    - 全站推广 → hybrid（拉新收割混合型）
    - 无法判断 → unknown（后续按转化率辅助判断）
    """
    name_lower = plan_name.lower()
    
    for kw in HARVEST_KEYWORDS:
        if kw in name_lower:
            return 'harvest'
    
    for kw in LAUNCH_KEYWORDS:
        if kw in name_lower:
            return 'launch'
    
    # 全站推广标记为混合型
    for kw in HYBRID_PLANS:
        if kw in name_lower:
            return 'hybrid'
    
    return 'unknown'


# ============================================================
# 二、推广报表解析
# ============================================================

class AdReportParser:
    """解析阿里妈妈推广报表"""
    
    def __init__(self, file_path: str, product_prefix: str = None):
        """
        Args:
            file_path: 推广报表xlsx文件路径
            product_prefix: 商品名前缀，用于筛选特定商品的计划（如'CG104'）
        """
        self.file_path = file_path
        self.product_prefix = product_prefix
        self.raw_rows = []
        self.plans = []  # 解析后的计划列表
        
    def parse(self) -> dict:
        """解析推广报表，返回结构化数据"""
        if not os.path.exists(self.file_path):
            return {'error': f'文件不存在: {self.file_path}'}
        
        # 读取文件
        rows = self._read_file()
        if not rows:
            return {'error': '无法读取文件或文件为空'}
        
        # 按商品前缀筛选
        if self.product_prefix:
            rows = [r for r in rows if self.product_prefix in str(r.get('plan_name', ''))]
        
        if not rows:
            return {'error': f'未找到包含"{self.product_prefix}"的计划'}
        
        # 解析每个计划
        self.plans = []
        for row in rows:
            plan = self._parse_plan(row)
            if plan:
                self.plans.append(plan)
        
        # 聚合分析
        return self._analyze()
    
    def _read_file(self) -> list:
        """读取xlsx/xls文件，返回行列表"""
        ext = os.path.splitext(self.file_path)[1].lower()
        
        if ext in ('.xlsx', '.xlsm') and openpyxl:
            return self._read_xlsx()
        elif ext == '.xls' and xlrd:
            return self._read_xls()
        elif ext == '.xlsx':
            return {'error': '需要openpyxl库来解析xlsx文件'}
        elif ext == '.xls':
            return {'error': '需要xlrd库来解析xls文件'}
        else:
            return []
    
    def _read_xlsx(self) -> list:
        """用openpyxl读取xlsx"""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            
            rows_data = []
            header = None
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # 跳过第一行（通常是汇总或标题行）
                
                # 找到表头行（包含"计划名称"等关键列名的行）
                if header is None:
                    row_values = [str(v).strip() if v else '' for v in row]
                    # 检查是否是表头行
                    if any('计划' in v and '名' in v for v in row_values):
                        header = row_values
                        continue
                    # 某些报表没有明确的表头标记，尝试用列号映射
                    continue
                
                # 数据行
                if row:
                    row_dict = {}
                    for j, val in enumerate(row):
                        if j < len(header):
                            row_dict[header[j]] = val
                    rows_data.append(row_dict)
            
            wb.close()
            
            # 如果没找到表头，尝试按列号映射
            if header is None:
                rows_data = self._read_xlsx_by_column()
            
            return rows_data
            
        except Exception as e:
            return []
    
    def _read_xlsx_by_column(self) -> list:
        """按固定列号映射读取（兜底方案）"""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            
            rows_data = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # 跳过首行
                
                if not row or len(row) < 10:
                    continue
                
                # 固定列号映射（阿里妈妈推广报表标准格式）
                row_dict = {
                    'promotion_type': row[5] if len(row) > 5 else None,   # 推广方式
                    'plan_name': row[6] if len(row) > 6 else None,        # 计划名称
                    'clicks': row[8] if len(row) > 8 else None,           # 点击量
                    'cost': row[9] if len(row) > 9 else None,             # 花费
                    'cpc': row[11] if len(row) > 11 else None,            # CPC
                    'direct_sales': row[19] if len(row) > 19 else None,   # 直接成交金额
                    'indirect_sales': row[20] if len(row) > 20 else None, # 间接成交金额
                    'total_sales': row[21] if len(row) > 21 else None,    # 总成交金额
                    'total_orders': row[22] if len(row) > 22 else None,   # 总成交笔数
                    'conv_rate': row[25] if len(row) > 25 else None,      # 点击转化率
                    'roi': row[26] if len(row) > 26 else None,            # 投入产出比
                    'cart_count': row[29] if len(row) > 29 else None,     # 总购物车数
                    'fav_cart_count': row[36] if len(row) > 36 else None, # 总收藏加购数
                    'fav_cart_cost': row[37] if len(row) > 37 else None,  # 总收藏加购成本
                    'cart_cost': row[43] if len(row) > 43 else None,      # 加购成本
                    'new_cust_count': row[61] if len(row) > 61 else None, # 成交新客数
                    'new_cust_ratio': row[62] if len(row) > 62 else None, # 成交新客占比
                    'cust_count': row[66] if len(row) > 66 else None,     # 成交人数
                    'natural_sales': row[69] if len(row) > 69 else None,  # 自然流量转化金额
                }
                rows_data.append(row_dict)
            
            wb.close()
            return rows_data
            
        except Exception as e:
            return []
    
    def _read_xls(self) -> list:
        """用xlrd读取xls（兜底）"""
        # 暂不实现，推广报表通常是xlsx格式
        return []
    
    def _safe_float(self, val, default=None) -> Optional[float]:
        """安全转float"""
        if val is None:
            return default
        try:
            if isinstance(val, str):
                val = val.replace(',', '').replace('%', '').strip()
                if val in ('-', '', '—', 'N/A'):
                    return default
            return float(val)
        except (ValueError, TypeError):
            return default
    
    def _parse_plan(self, row: dict) -> Optional[dict]:
        """解析单行计划数据"""
        # 支持列名映射和列号映射两种格式
        plan_name = str(row.get('plan_name', row.get('计划名称', ''))).strip()
        if not plan_name:
            return None
        
        # 尝试从列名或列号映射取值
        clicks = self._safe_float(row.get('clicks', row.get('点击量')))
        cost = self._safe_float(row.get('cost', row.get('花费')))
        cpc = self._safe_float(row.get('cpc', row.get('CPC')))
        total_sales = self._safe_float(row.get('total_sales', row.get('总成交金额')))
        total_orders = self._safe_float(row.get('total_orders', row.get('总成交笔数')))
        conv_rate = self._safe_float(row.get('conv_rate', row.get('点击转化率')))
        roi = self._safe_float(row.get('roi', row.get('投入产出比')))
        cart_count = self._safe_float(row.get('cart_count', row.get('总购物车数')))
        fav_cart_count = self._safe_float(row.get('fav_cart_count', row.get('总收藏加购数')))
        fav_cart_cost = self._safe_float(row.get('fav_cart_cost', row.get('总收藏加购成本')))
        cart_cost = self._safe_float(row.get('cart_cost', row.get('加购成本')))
        new_cust_ratio = self._safe_float(row.get('new_cust_ratio', row.get('成交新客占比')))
        natural_sales = self._safe_float(row.get('natural_sales', row.get('自然流量转化金额')))
        new_cust_count = self._safe_float(row.get('new_cust_count', row.get('成交新客数')))
        cust_count = self._safe_float(row.get('cust_count', row.get('成交人数')))
        
        # 计算加购率
        cart_rate = None
        if clicks and clicks > 0 and fav_cart_count and fav_cart_count > 0:
            cart_rate = round(fav_cart_count / clicks * 100, 2)
        
        # 归因修正：计算ROI
        surface_roi = None
        if cost and cost > 0 and total_sales and total_sales > 0:
            effective_sales = total_sales
            if natural_sales and natural_sales > 0:
                effective_sales = total_sales - natural_sales
            surface_roi = round(effective_sales / cost, 2)
        
        # 归因虚高比例
        natural_ratio = None
        if total_sales and total_sales > 0 and natural_sales and natural_sales > 0:
            natural_ratio = round(natural_sales / total_sales * 100, 1)
        
        # 计划类型
        plan_type = classify_plan(plan_name)
        
        return {
            'plan_name': plan_name,
            'plan_type': plan_type,  # launch/harvest/special/unknown
            'clicks': clicks,
            'cost': cost,
            'cpc': cpc,
            'total_sales': total_sales,
            'total_orders': total_orders,
            'conv_rate': conv_rate,         # 点击转化率(%)
            'surface_roi': roi,             # 表面ROI

            'cart_count': cart_count,
            'fav_cart_count': fav_cart_count,
            'fav_cart_cost': fav_cart_cost,  # 收藏加购成本
            'cart_cost': cart_cost,          # 加购成本
            'cart_rate': cart_rate,          # 加购率(%)
            'new_cust_ratio': new_cust_ratio, # 成交新客占比(%)
            'new_cust_count': new_cust_count,  # 成交新客数
            'cust_count': cust_count,          # 成交人数
            'natural_sales': natural_sales,
            'natural_ratio': natural_ratio,   # 自然流量转化占比(%)
        }
    
    def _analyze(self) -> dict:
        """聚合分析所有计划——先按计划名合并每日数据，再做诊断"""
        if not self.plans:
            return {'error': '没有有效的计划数据'}
        
        # ===== 关键：按计划名聚合每日数据 =====
        # 推广报表是每天每计划一行，需要合并成单个计划
        merged = defaultdict(lambda: {
            'plan_name': '',
            'plan_type': 'unknown',
            'clicks': 0,
            'cost': 0,
            'total_sales': 0,
            'total_orders': 0,
            'cart_count': 0,
            'fav_cart_count': 0,
            'fav_cart_cost_total': 0,  # 加权用
            'cart_cost_total': 0,       # 加权用
            'natural_sales': 0,
            'new_cust_count': 0,
            'cust_count': 0,
            'cpc_samples': [],    # CPC加权用
            'conv_rate_samples': [],  # 转化率加权用
            'new_cust_ratio_samples': [],  # 新客占比加权用
            'day_count': 0,
        })
        
        for p in self.plans:
            name = p['plan_name']
            m = merged[name]
            m['plan_name'] = name
            m['plan_type'] = p['plan_type']
            m['day_count'] += 1
            
            clicks = p.get('clicks') or 0
            cost = p.get('cost') or 0
            m['clicks'] += clicks
            m['cost'] += cost
            m['total_sales'] += p.get('total_sales') or 0
            m['total_orders'] += p.get('total_orders') or 0
            m['cart_count'] += p.get('cart_count') or 0
            m['fav_cart_count'] += p.get('fav_cart_count') or 0
            m['natural_sales'] += p.get('natural_sales') or 0
            m['new_cust_count'] += p.get('new_cust_count') or 0
            m['cust_count'] += p.get('cust_count') or 0
            
            # 加权汇总：收藏加购成本和加购成本
            if p.get('fav_cart_cost') and p.get('fav_cart_count'):
                m['fav_cart_cost_total'] += p['fav_cart_cost'] * p['fav_cart_count']
            if p.get('cart_cost') and p.get('cart_count'):
                m['cart_cost_total'] += p['cart_cost'] * p['cart_count']
        
        # 计算合并后的指标
        self.plans = []
        for name, m in merged.items():
            clicks = m['clicks']
            cost = m['cost']
            total_sales = m['total_sales']
            natural_sales = m['natural_sales']
            fav_cart_count = m['fav_cart_count']
            
            # 加购率
            cart_rate = round(fav_cart_count / clicks * 100, 2) if clicks > 0 else None
            
            # ROI
            effective_sales = total_sales - natural_sales
            surface_roi = round(effective_sales / cost, 2) if cost > 0 else None
            
            # 表面ROI
            surface_roi = round(total_sales / cost, 2) if cost > 0 else None
            
            # 自然占比
            natural_ratio = round(natural_sales / total_sales * 100, 1) if total_sales > 0 else 0
            
            # 转化率
            conv_rate = round(m['total_orders'] / clicks * 100, 2) if clicks > 0 else 0
            
            # CPC
            cpc = round(cost / clicks, 2) if clicks > 0 else None
            
            # 收藏加购成本（加权平均）
            fav_cart_cost = round(m['fav_cart_cost_total'] / fav_cart_count, 2) if fav_cart_count > 0 else None
            
            # 加购成本（加权平均）
            cart_cost = round(m['cart_cost_total'] / m['cart_count'], 2) if m['cart_count'] > 0 else None
            
            # 新客占比（加权平均）
            new_cust_ratio = round(m['new_cust_count'] / m['cust_count'] * 100, 1) if m['cust_count'] > 0 else None
            
            self.plans.append({
                'plan_name': name,
                'plan_type': m['plan_type'],
                'clicks': clicks,
                'cost': round(cost, 2),
                'cpc': cpc,
                'total_sales': round(total_sales, 2),
                'total_orders': m['total_orders'],
                'conv_rate': conv_rate,
                'surface_roi': surface_roi,
                'surface_roi': surface_roi,
                'cart_count': m['cart_count'],
                'fav_cart_count': fav_cart_count,
                'fav_cart_cost': fav_cart_cost,
                'cart_cost': cart_cost,
                'cart_rate': cart_rate,
                'new_cust_ratio': new_cust_ratio,
                'natural_sales': round(natural_sales, 2),
                'natural_ratio': natural_ratio,
                'day_count': m['day_count'],
            })
        
        # ===== 以下为原有聚合逻辑（基于合并后的plans） =====
        # 按推广方式和拉新/收割分组
        by_type = defaultdict(list)
        by_promotion = defaultdict(list)
        
        for plan in self.plans:
            by_type[plan['plan_type']].append(plan)
            promo = self._extract_promotion_type(plan['plan_name'])
            by_promotion[promo].append(plan)
        
        # 汇总统计
        total_cost = sum(p['cost'] or 0 for p in self.plans)
        total_sales = sum(p['total_sales'] or 0 for p in self.plans)
        total_natural = sum(p['natural_sales'] or 0 for p in self.plans)
        total_clicks = sum(p['clicks'] or 0 for p in self.plans)
        
        # ===== 混合型计划拆分（货品全站推广等） =====
        hybrid_ratio = (self.ctx if hasattr(self, 'ctx') else None)
        launch_ratio = DEFAULT_HYBRID_LAUNCH_RATIO
        # 兼容：如果调用方传了hybrid_launch_ratio就用
        if hasattr(self, '_hybrid_launch_ratio'):
            launch_ratio = self._hybrid_launch_ratio
        
        hybrid_plans = by_type.get('hybrid', [])
        hybrid_launch_plans = []
        hybrid_harvest_plans = []
        
        for p in hybrid_plans:
            # 按比例拆分花费和成交
            p_launch = dict(p)
            p_launch['plan_type'] = 'launch'
            p_launch['cost'] = round((p.get('cost') or 0) * launch_ratio, 2)
            p_launch['total_sales'] = round((p.get('total_sales') or 0) * launch_ratio, 2)
            p_launch['total_orders'] = round((p.get('total_orders') or 0) * launch_ratio, 0)
            p_launch['natural_sales'] = round((p.get('natural_sales') or 0) * launch_ratio, 2)
            p_launch['cart_count'] = round((p.get('cart_count') or 0) * launch_ratio, 0)
            p_launch['fav_cart_count'] = round((p.get('fav_cart_count') or 0) * launch_ratio, 0)
            p_launch['new_cust_count'] = round((p.get('new_cust_count') or 0) * launch_ratio, 0)
            p_launch['cust_count'] = round((p.get('cust_count') or 0) * launch_ratio, 0)
            p_launch['clicks'] = round((p.get('clicks') or 0) * launch_ratio, 0)
            # 重新计算比率类指标
            if p_launch['clicks'] and p_launch['clicks'] > 0 and p_launch.get('fav_cart_count') and p_launch['fav_cart_count'] > 0:
                p_launch['cart_rate'] = round(p_launch['fav_cart_count'] / p_launch['clicks'] * 100, 2)
            if p_launch['cost'] and p_launch['cost'] > 0 and p_launch.get('total_sales') and p_launch['total_sales'] > 0:
                p_launch['surface_roi'] = round(p_launch['total_sales'] / p_launch['cost'], 2)
                effective = p_launch['total_sales'] - (p_launch.get('natural_sales') or 0)
                p_launch['surface_roi'] = round(effective / p_launch['cost'], 2)
            if p_launch['clicks'] and p_launch['clicks'] > 0 and p_launch.get('total_orders'):
                p_launch['conv_rate'] = round(p_launch['total_orders'] / p_launch['clicks'] * 100, 2)
            if p_launch['cost'] and p_launch['cost'] > 0 and p_launch.get('fav_cart_count') and p_launch['fav_cart_count'] > 0:
                p_launch['fav_cart_cost'] = round(p_launch['cost'] / p_launch['fav_cart_count'], 2)
            p_launch['plan_name'] = p.get('plan_name', '') + '【拉新部分】'
            hybrid_launch_plans.append(p_launch)
            
            p_harvest = dict(p)
            p_harvest['plan_type'] = 'harvest'
            harvest_ratio = 1 - launch_ratio
            p_harvest['cost'] = round((p.get('cost') or 0) * harvest_ratio, 2)
            p_harvest['total_sales'] = round((p.get('total_sales') or 0) * harvest_ratio, 2)
            p_harvest['total_orders'] = round((p.get('total_orders') or 0) * harvest_ratio, 0)
            p_harvest['natural_sales'] = round((p.get('natural_sales') or 0) * harvest_ratio, 2)
            p_harvest['cart_count'] = round((p.get('cart_count') or 0) * harvest_ratio, 0)
            p_harvest['fav_cart_count'] = round((p.get('fav_cart_count') or 0) * harvest_ratio, 0)
            p_harvest['new_cust_count'] = round((p.get('new_cust_count') or 0) * harvest_ratio, 0)
            p_harvest['cust_count'] = round((p.get('cust_count') or 0) * harvest_ratio, 0)
            p_harvest['clicks'] = round((p.get('clicks') or 0) * harvest_ratio, 0)
            # 重新计算比率类指标
            if p_harvest['clicks'] and p_harvest['clicks'] > 0 and p_harvest.get('fav_cart_count') and p_harvest['fav_cart_count'] > 0:
                p_harvest['cart_rate'] = round(p_harvest['fav_cart_count'] / p_harvest['clicks'] * 100, 2)
            if p_harvest['cost'] and p_harvest['cost'] > 0 and p_harvest.get('total_sales') and p_harvest['total_sales'] > 0:
                p_harvest['surface_roi'] = round(p_harvest['total_sales'] / p_harvest['cost'], 2)
                effective = p_harvest['total_sales'] - (p_harvest.get('natural_sales') or 0)
                p_harvest['surface_roi'] = round(effective / p_harvest['cost'], 2)
            if p_harvest['clicks'] and p_harvest['clicks'] > 0 and p_harvest.get('total_orders'):
                p_harvest['conv_rate'] = round(p_harvest['total_orders'] / p_harvest['clicks'] * 100, 2)
            if p_harvest['cost'] and p_harvest['cost'] > 0 and p_harvest.get('fav_cart_count') and p_harvest['fav_cart_count'] > 0:
                p_harvest['fav_cart_cost'] = round(p_harvest['cost'] / p_harvest['fav_cart_count'], 2)
            p_harvest['plan_name'] = p.get('plan_name', '') + '【收割部分】'
            hybrid_harvest_plans.append(p_harvest)
        
        # ===== 拉新层：只看纯拉新计划，不含混合型计划的拆分 =====
        # 混合型计划（全站推广）的1:1拆分是人为分配，不能代表真实拉新效果
        launch_plans = by_type.get('launch', []) + by_type.get('unknown', [])
        launch_cost = sum(p['cost'] or 0 for p in launch_plans)
        launch_sales = sum(p['total_sales'] or 0 for p in launch_plans)
        launch_natural = sum(p['natural_sales'] or 0 for p in launch_plans)
        launch_surface_roi = round(launch_sales / launch_cost, 2) if launch_cost > 0 else None
        
        # ===== 收割层：只看纯收割计划，不含混合型计划的拆分 =====
        harvest_plans = by_type.get('harvest', [])
        harvest_cost = sum(p['cost'] or 0 for p in harvest_plans)
        harvest_sales = sum(p['total_sales'] or 0 for p in harvest_plans)
        harvest_natural = sum(p['natural_sales'] or 0 for p in harvest_plans)
        harvest_surface_roi = round(harvest_sales / harvest_cost, 2) if harvest_cost > 0 else None
        
        # ===== 混合型计划（全站推广）：单独统计，不拆分 =====
        hybrid_cost = sum(p['cost'] or 0 for p in hybrid_plans)
        hybrid_sales = sum(p['total_sales'] or 0 for p in hybrid_plans)
        hybrid_natural = sum(p['natural_sales'] or 0 for p in hybrid_plans)
        hybrid_surface_roi = round(hybrid_sales / hybrid_cost, 2) if hybrid_cost > 0 else None
        
        return {
            # 全局汇总
            'total_cost': round(total_cost, 2),
            'total_sales': round(total_sales, 2),
            'total_natural': round(total_natural, 2),
            'total_clicks': total_clicks,
            'surface_roi': round(total_sales / total_cost, 2) if total_cost > 0 else None,

            'natural_attribution_ratio': round(total_natural / total_sales * 100, 1) if total_sales > 0 else 0,
            'plan_count': len(self.plans),
            
            # 混合型计划信息
            'hybrid_plans': [{'plan_name': p.get('plan_name', ''), 'cost': p.get('cost', 0), 'launch_ratio': launch_ratio}],
            
            # 拉新层（只含纯拉新计划，不含混合型拆分）
            'launch': {
                'cost': round(launch_cost, 2),
                'sales': round(launch_sales, 2),
                'natural': round(launch_natural, 2),
                'surface_roi': launch_surface_roi,
                'plans': launch_plans,
            },
            
            # 收割层（只含纯收割计划，不含混合型拆分）
            'harvest': {
                'cost': round(harvest_cost, 2),
                'sales': round(harvest_sales, 2),
                'natural': round(harvest_natural, 2),
                'surface_roi': harvest_surface_roi,
                'plans': harvest_plans,
            },
            
            # 混合型计划层（全站推广等，单独统计不拆分）
            'hybrid': {
                'cost': round(hybrid_cost, 2),
                'sales': round(hybrid_sales, 2),
                'natural': round(hybrid_natural, 2),
                'surface_roi': hybrid_surface_roi,
                'plans': hybrid_plans,
            },
            
            # 按推广方式分组
            'by_promotion': dict(by_promotion),
            
            # 所有计划明细（合并后，含原始hybrid类型）
            'all_plans': self.plans,
        }
    
    def _extract_promotion_type(self, plan_name: str) -> str:
        """从计划名提取推广方式
        
        常见格式：
        - 自定义出价-CG104-xxx
        - 关键词精准-CG104-xxx
        - 超级短视频-CG104-xxx
        - 好货快投成交转化-CG104-xxx
        """
        parts = plan_name.split('-')
        if len(parts) >= 1:
            return parts[0].strip()
        return plan_name


# ============================================================
# 三、推广诊断结论生成
# ============================================================

class AdDiagnosisGenerator:
    """基于推广数据生成深度诊断结论"""
    
    def __init__(self, ad_data: dict, product_context: dict = None):
        """
        Args:
            ad_data: AdReportParser.parse() 的返回结果
            product_context: 商品基本数据（自然转化率、客单价等）
        """
        self.ad = ad_data
        self.ctx = product_context or {}
    
    def generate(self) -> dict:
        """生成推广诊断结论"""
        if 'error' in self.ad:
            return {'error': self.ad['error']}
        
        # 先计算各层诊断
        launch_diagnosis = self._diagnose_launch()
        harvest_diagnosis = self._diagnose_harvest()
        attribution_warning = self._generate_attribution_warning()
        refund_impact = self._estimate_refund_impact()
        
        # 把诊断结果缓存到self上，供_generate_actions使用
        self._launch_diagnosis = launch_diagnosis
        self._harvest_diagnosis = harvest_diagnosis
        
        return {
            'summary': self._generate_summary(),
            'launch_diagnosis': launch_diagnosis,
            'harvest_diagnosis': harvest_diagnosis,
            'attribution_warning': attribution_warning,
            'refund_impact': refund_impact,
            'actions': self._generate_actions(),
            # 透传hybrid层数据（全站推广等混合型计划，不拆分）
            'hybrid': self.ad.get('hybrid', {}),
        }
    
    def _generate_summary(self) -> dict:
        """推广整体诊断摘要"""
        total_cost = self.ad.get('total_cost', 0)
        surface_roi = self.ad.get('surface_roi')
        natural_ratio = self.ad.get('natural_attribution_ratio', 0)
        
        # ROI判断
        roi_level = 'unknown'
        if surface_roi is not None:
            if surface_roi >= 3.0:
                roi_level = 'excellent'
            elif surface_roi >= 2.0:
                roi_level = 'ok'
            elif surface_roi >= 1.0:
                roi_level = 'weak'
            else:
                roi_level = 'loss'
        
        # 归因虚高程度
        inflation_level = 'none'
        if natural_ratio > 25:
            inflation_level = 'severe'   # 严重虚高
        elif natural_ratio > 15:
            inflation_level = 'moderate' # 中度
        elif natural_ratio > 5:
            inflation_level = 'mild'     # 轻度
        
        return {
            'total_cost': total_cost,
            'surface_roi': surface_roi,
            'roi_level': roi_level,
            'roi_inflation': 0,  # Deprecated: always 0 with platform ROI
            'inflation_level': inflation_level,
            'natural_attribution_ratio': natural_ratio,
        }
    
    def _diagnose_launch(self) -> dict:
        """拉新层诊断
        
        评价标准：蓄水效率（加购率+加购成本+收藏加购成本）
        不看ROI——拉新本来就不指望直接成交
        """
        launch = self.ad.get('launch', {})
        plans = launch.get('plans', [])
        
        if not plans:
            return {'status': 'no_data', 'message': '无拉新计划数据'}
        
        # 找最优拉新渠道
        best_cart_rate = None
        best_cart_plan = None
        worst_plans = []
        
        for p in plans:
            if p.get('cost') and p['cost'] > 0:
                # 蓄水效率评估
                cart_rate = p.get('cart_rate')
                fav_cart_cost = p.get('fav_cart_cost')
                surface_roi = p.get('surface_roi')
                
                # 综合蓄水得分（加购率权重60% + 成本效率40%）
                score = 0
                if cart_rate is not None:
                    score += min(cart_rate / 20 * 6, 6)  # 加购率20%满分6分
                if fav_cart_cost is not None and fav_cart_cost > 0:
                    score += min(20 / fav_cart_cost * 4, 4)  # 收藏加购成本越低越好，¥5满分4分
                
                p['launch_score'] = round(score, 1)
                
                if best_cart_rate is None or (cart_rate is not None and cart_rate > (best_cart_rate or 0)):
                    best_cart_rate = cart_rate
                    best_cart_plan = p
                
                # 标记低效拉新：蓄水得分<4 或 蓄水得分4-6但ROI<1（蓄水转化都不行）
                is_low_efficiency = False
                if score < 4:
                    is_low_efficiency = True
                elif score < 6 and (surface_roi is None or surface_roi < 1.0):
                    is_low_efficiency = True
                
                if is_low_efficiency:
                    worst_plans.append(p)
        
        # 按蓄水效率排序
        scored_plans = sorted(plans, key=lambda x: x.get('launch_score', 0), reverse=True)
        
        return {
            'total_cost': launch.get('cost'),
            'total_sales': launch.get('sales'),
            'surface_roi': launch.get('surface_roi'),
            'best_cart_plan': best_cart_plan,
            'worst_plans': worst_plans,
            'ranked_plans': scored_plans,
        }
    
    def _diagnose_harvest(self) -> dict:
        """收割层诊断
        
        评价标准：ROI必须高，打的是暖客/加购/收藏人群
        ROI<1就是真浪费，转化率应>3%
        """
        harvest = self.ad.get('harvest', {})
        plans = harvest.get('plans', [])
        
        if not plans:
            return {'status': 'no_data', 'message': '无收割计划数据'}
        
        qualified_plans = []
        weak_plans = []
        waste_plans = []
        
        for p in plans:
            if p.get('cost') and p['cost'] > 0:
                surface_roi = p.get('surface_roi')
                conv_rate = p.get('conv_rate')
                
                # 收割效率评估
                if surface_roi is not None and (surface_roi >= 3.0 or (surface_roi >= 2.5 and conv_rate is not None and conv_rate >= 1.0)):
                    p['harvest_level'] = 'qualified'
                    qualified_plans.append(p)
                elif surface_roi is not None and surface_roi >= 1.5:
                    p['harvest_level'] = 'weak'
                    weak_plans.append(p)
                else:
                    p['harvest_level'] = 'waste'
                    waste_plans.append(p)
                
                # 成交新客占比解读（不能用于判断是否在拉新）
                new_cust_ratio = p.get('new_cust_ratio')
                if new_cust_ratio is not None:
                    p['new_cust_note'] = (
                        f'成交新客占比{new_cust_ratio}%属于正常现象'
                        f'（阿里妈妈"成交新客"=未在本店购买过的客户，'
                        f'收割计划促首单转化，新客占比高是正常的）'
                    )
        
        return {
            'total_cost': harvest.get('cost'),
            'total_sales': harvest.get('sales'),
            'surface_roi': harvest.get('surface_roi'),
            'qualified_plans': qualified_plans,
            'weak_plans': weak_plans,
            'waste_plans': waste_plans,
        }
    
    def _generate_attribution_warning(self) -> list:
        """归因虚高警告"""
        warnings = []
        
        for p in self.ad.get('all_plans', []):
            natural_ratio = p.get('natural_ratio')
            if natural_ratio is not None and natural_ratio > 20:
                warnings.append({
                    'plan': p['plan_name'],
                    'type': 'severe_attribution_inflation',
                    'message': f'{p["plan_name"]}：自然流量转化占比{natural_ratio}%，近归因有一定虚高{p.get("surface_roi", "?")}',
                    'natural_ratio': natural_ratio,
                })
            elif natural_ratio is not None and natural_ratio > 10:
                warnings.append({
                    'plan': p['plan_name'],
                    'type': 'moderate_attribution_inflation',
                    'message': f'{p["plan_name"]}：自然流量转化占比{natural_ratio}%，ROI有一定虚高',
                    'natural_ratio': natural_ratio,
                })
        
        return warnings
    
    def _estimate_refund_impact(self) -> dict:
        """退款/秒退独立诊断
        
        退款是产品/服务维度的问题，不影响推广ROI评价
        （推广ROI只看ROI=剥离自然归因后的ROI）
        这里只做退款根因分析，供产品/服务维度引用
        """
        refund_rate = self.ctx.get('refund_rate')
        is_instant_refund = self.ctx.get('instant_refund', False)
        
        if refund_rate is None:
            return {'status': 'no_data', 'message': '缺少退款率数据'}
        
        result = {
            'refund_rate': refund_rate,
            'is_instant_refund': is_instant_refund,
        }
        
        # 退款严重程度
        if refund_rate >= 40:
            result['severity'] = 'severe'
        elif refund_rate >= 25:
            result['severity'] = 'moderate'
        else:
            result['severity'] = 'normal'
        
        # 秒退根因判断（V6.1：全站推广凑单场景识别）
        if is_instant_refund:
            ad_ratio = self.ctx.get('ad_traffic_ratio', 0)
            has_hybrid_plans = bool(self.ad.get('hybrid_plans'))
            
            if has_hybrid_plans:
                # 货品全站推广混合型计划 → 凑单退款是主因
                result['refund_root_cause'] = 'hybrid_plan_cart_stuffing'
                result['refund_explanation'] = (
                    f'秒退为主+全站推广计划在跑，根因是全站推广的凑单行为——'
                    f'用户为了凑满减/包邮把商品加进购物车，付款后立刻退款，'
                    f'从来就没打算买这个商品。这不是流量不精准或产品问题，'
                    f'而是全站推广场景的特殊行为模式。'
                    f'注意：退款不影响推广ROI评价，但影响实际净收入，需单独关注'
                )
            elif ad_ratio > 0.5:
                result['refund_root_cause'] = 'traffic_imprecision'
                result['refund_explanation'] = (
                    f'秒退为主+广告占比{int(ad_ratio*100)}%，根因是流量不精准导致的虚假成交，'
                    f'非产品/售后问题。广告转化率越低，秒退比例通常越高。'
                    f'注意：退款是流量质量问题在成交端的体现，需结合推广精准度一起看'
                )
            else:
                result['refund_root_cause'] = 'need_more_data'
                result['refund_explanation'] = '秒退为主，但广告占比不高且无全站推广，需进一步排查退款渠道分布'
        else:
            result['refund_root_cause'] = 'unknown'
            result['refund_explanation'] = '退款原因未明确，需退款明细数据'
        
        return result
    
    def _generate_actions(self) -> list:
        """生成推广优化动作"""
        actions = []
        
        # 1. 砍掉真浪费的计划
        waste_plans = self._harvest_diagnosis.get('waste_plans', [])
        for p in waste_plans:
            actions.append({
                'priority': 'urgent',
                'action': 'kill',
                'target': p['plan_name'],
                'reason': f'ROI仅{p.get("surface_roi", "?")}{f"，CPC ¥{p.get("cpc", "?")}" if p.get("cpc") else ""}，无蓄水价值，必须砍',
                'save_cost': p.get('cost'),
            })
        
        # 2. 减预算低效拉新
        worst_plans = self._launch_diagnosis.get('worst_plans', [])
        for p in worst_plans:
            surface_roi = p.get('surface_roi')
            natural_ratio = p.get('natural_ratio', 0)
            save_pct = 50 if natural_ratio > 20 else 30
            
            actions.append({
                'priority': 'important',
                'action': 'reduce_budget',
                'target': p['plan_name'],
                'reason': (
                    f'蓄水效率低（得分{p.get("launch_score", "?")}），'
                    f'ROI {surface_roi if surface_roi else "数据暂缺（基于蓄水效率判断，建议在无界后台确认实时ROI）"}'
                    ''
                ),
                'save_cost': round(p.get('cost', 0) * save_pct / 100, 2),
                'reduce_pct': save_pct,
            })
        
        # 3. 加量最优拉新
        best_plan = self._launch_diagnosis.get('best_cart_plan')
        if best_plan:
            cart_rate = best_plan.get('cart_rate')
            fav_cart_cost = best_plan.get('fav_cart_cost')
            actions.append({
                'priority': 'important',
                'action': 'increase_budget',
                'target': best_plan['plan_name'],
                'reason': (
                    f'蓄水效率全场最高'
                    f'{f"（加购率{cart_rate}%，收藏加购成本¥{fav_cart_cost}" if cart_rate and fav_cart_cost else ""}），'
                    f'ROI {best_plan.get("surface_roi", "?")}也健康，加量是最优选择'
                ),
                'current_cost': best_plan.get('cost'),
            })
        
        # 4. 检查收割人群包（转化率偏低的收割计划）
        weak_plans = self._harvest_diagnosis.get('weak_plans', [])
        for p in weak_plans:
            conv_rate = p.get('conv_rate')
            if conv_rate is not None and conv_rate < 3.0:
                actions.append({
                    'priority': 'normal',
                    'action': 'check_audience',
                    'target': p['plan_name'],
                    'reason': (
                        f'作为收割计划转化率仅{conv_rate}%（应>3%），'
                        f'人群精准度可能不够，需回无界后台确认人群包是'
                        f'"已加购/收藏未购"还是泛化的"相似人群"'
                    ),
                })
        
        return actions


# ============================================================
# 四、便捷函数
# ============================================================

def parse_ad_report(file_path: str, product_prefix: str = None) -> dict:
    """便捷函数：解析推广报表并生成诊断"""
    parser = AdReportParser(file_path, product_prefix)
    return parser.parse()


def diagnose_ad(file_path: str, product_prefix: str = None, product_context: dict = None) -> dict:
    """便捷函数：解析推广报表 + 生成推广诊断"""
    ad_data = parse_ad_report(file_path, product_prefix)
    if 'error' in ad_data:
        return ad_data
    
    generator = AdDiagnosisGenerator(ad_data, product_context)
    return generator.generate()
